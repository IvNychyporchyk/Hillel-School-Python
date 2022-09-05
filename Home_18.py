# Прочитать сохранённый csv-файл из задания №17 и сохранить данные
# в excel-файл, кроме возраста – столбец с этими данными не нужен.

import openpyxl
import csv

# with open('home17.csv', encoding='utf-8') as f:
#     file_read = csv.reader(f)
#     count = 0
#     for row in file_read:
#         print(f'{row[0]} | {row[1]} | {row[2]}')
#     if count == 0:
#         print('-' * 20)
#     count += 1


work_book = openpyxl.Workbook()
print(work_book.sheetnames)
work_sheet = work_book.create_sheet(title='Home sheet', index=0)
with open('home17.csv', encoding='utf-8') as f:
    csv_file = list(csv.reader(f))

for row_index, row in enumerate(csv_file):
    for col_index, value in enumerate(row):
        cell = work_sheet.cell(row=col_index + 1, column=row_index + 1)
        cell.value = value
work_sheet.delete_rows(2)
work_sheet.insert_rows(1)
work_sheet['B1'] = 'Person_1'
work_sheet['C1'] = 'Person_2'
work_sheet['D1'] = 'Person_3'
work_sheet['E1'] = 'Person_4'
work_sheet['F1'] = 'Person_5'
work_sheet.insert_rows(2)
work_sheet['A2'] = 'id_'
work_sheet['B2'] = '678945'
work_sheet['C2'] = '457328'
work_sheet['D2'] = '897678'
work_sheet['E2'] = '978575'
work_sheet['F2'] = '788765'

work_book.save("home18.xlsx")
# work_book = openpyxl.load_workbook("home18.xlsx")


work_b = openpyxl.load_workbook('home18.xlsx')

w_sheet = work_b['Home sheet']

cell_ = w_sheet.cell(row=1, column=1)
print(cell_)

# name_of_filds = ['Name', 'Grade', 'Age']
# field_01 = ['Jack', '3', '10']
# field_02 = ['Tom', '5', '12']
# field_03 = ['Sam', '11', '18']
#
#
# wb = openpyxl.Workbook()
# print(wb.sheetnames)
#
# wb.create_sheet(title='Первый лист', index=0)
# print(wb.sheetnames)
#
# wb.remove(wb['Sheet'])
# print(wb.sheetnames)
#
# sheet = wb['Первый лист']
# print(sheet)
#
# for row_index, row in enumerate((name_of_filds, field_01, field_02, field_03)):
#     for col_index, value in enumerate(row):
#         cell = sheet.cell(row=row_index+1, column=col_index+1)
#         cell.value = value
#
# wb.save('task_03.xlsx')
#
#
# wb = openpyxl.load_workbook('task_03.xlsx')
# sheets = wb.sheetnames
# print(sheets)
#
# sheet_2 = wb.active
# print(sheet_2)
#
# sheet = wb['Первый лист']
# print(sheet)
#
# print(sheet['D1'].value)
#
# rows = sheet.max_row
# cols = sheet.max_column
#
# print(rows)
# print(cols)
#
# name_of_fields = []
# fields_values = []
#
# for i in range(1, rows+1):
#     value_row = []
#     for j in range(1, cols+1):
#         cell = sheet.cell(row=i, column=j)
#         value = str(cell.value)
#         if i == 1:
#             name_of_fields.append(value)
#         else:
#             value_row.append(value)
#     if i != 1:
#         fields_values.append(value_row)
#
# print(name_of_fields)
# print(fields_values)
